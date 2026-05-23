"""Excel import/export for standards — complete indicator fields."""
from typing import List
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from ..models.standard import StdCategory, StdIndicator, StdRequirement


def parse_standards_excel(file_path: str, db: Session) -> int:
    """Parse standards Excel and import into DB. Returns count of indicators added."""
    wb = load_workbook(file_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    category_cache = {}
    count = 0

    for row in rows:
        if not row or not row[0]:
            continue
        cat_name = str(row[0]).strip() if len(row) > 0 else ""
        sub_cat = str(row[1]).strip() if len(row) > 1 else ""
        code = str(row[2]).strip() if len(row) > 2 else ""
        name = str(row[3]).strip() if len(row) > 3 else ""
        std_val = str(row[4]).strip() if len(row) > 4 else ""
        unit = str(row[5]).strip() if len(row) > 5 else ""
        itype = str(row[6]).strip() if len(row) > 6 else "numeric_less_equal"
        weight_raw = str(row[7]).strip() if len(row) > 7 else "0"

        if not code or not name:
            continue

        # Category: use sub_cat if present, else cat_name
        cat_key = (cat_name, sub_cat) if sub_cat else (cat_name,)
        if cat_key not in category_cache:
            existing = db.query(StdCategory).filter(
                StdCategory.name == cat_name
            ).first()
            if not existing:
                existing = StdCategory(name=cat_name, code=cat_name[:20], weight=0, sort_order=len(category_cache) + 1)
                db.add(existing)
                db.flush()
            category_cache[cat_key] = existing.id

        cat_id = category_cache[cat_key]

        try:
            weight = float(weight_raw)
        except ValueError:
            weight = 0.0

        ind = StdIndicator(
            category_id=cat_id,
            code=code,
            name=name,
            standard_value=std_val,
            unit=unit,
            indicator_type=itype if itype in ('numeric_less_equal', 'numeric_greater_equal', 'numeric_equal', 'numeric_range', 'yesno') else 'numeric_less_equal',
            weight=weight,
            max_score=100,
        )
        db.add(ind)
        count += 1

    db.commit()
    return count


def generate_template() -> Workbook:
    """Generate an Excel template with headers and example rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "评审标准导入模板"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    example_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    # Sheet 1: Import template
    headers = ["分类", "子分类(可选)", "编码", "指标名称", "标准值", "单位", "判定类型", "权重"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Example data
    examples = [
        ["医疗质量-死亡类", "", "QL01", "住院患者总死亡率", "≤0.8%", "%", "numeric_less_equal", 5],
        ["医疗质量-死亡类", "", "QL02", "手术患者死亡率", "≤0.2%", "%", "numeric_less_equal", 5],
        ["资源配置-床位", "", "BD03", "重症医学科床位占比", "≥2%", "%", "numeric_greater_equal", 5],
        ["药事管理", "", "PH01", "处方合格率", "≥95%", "%", "numeric_greater_equal", 5],
        ["前置-依法执业", "", "PR06", "未出租承包科室", "是", "", "yesno", 0],
    ]
    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = example_fill
            cell.border = thin_border

    # Column widths
    widths = [20, 16, 10, 28, 14, 8, 24, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Sheet 2: Instructions
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["字段", "说明", "示例"],
        ["分类", "指标所属分类名称", "医疗质量-死亡类"],
        ["子分类", "二级分类（可选，留空则使用分类）", ""],
        ["编码", "指标唯一编码，如 QL01", "QL01"],
        ["指标名称", "指标中文全称", "住院患者总死亡率"],
        ["标准值", "达标判断标准，支持 ≤、≥、= 运算符", "≤0.8%"],
        ["单位", "指标单位: %, ‰, 张, 人次, 天 等", "%"],
        ["判定类型", "numeric_less_equal / numeric_greater_equal / numeric_equal / numeric_range / yesno", "numeric_less_equal"],
        ["权重", "整数 0-100，0 = 前置否决项不计入总分", "5"],
    ]
    for r, row_data in enumerate(instructions, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            cell.border = thin_border

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 30

    return wb
