from typing import List, Optional
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..database import get_db
from ..models.assessment import Assessment, AssessmentItem
from ..models.standard import StdCategory, StdIndicator
from ..models.standard_set import StandardSet
from ..models.department import Department
from ..models.audit_log import AuditLog
from ..middleware.tenant import get_current_tenant_id, get_current_user, get_current_user_tenant
from ..models.user import UserTenant
from ..services.compliance import check_compliance


def _log(db, user_id, tenant_id, action, target_type, target_id, detail=""):
    db.add(AuditLog(user_id=user_id, tenant_id=tenant_id, action=action,
                     target_type=target_type, target_id=target_id, detail=detail))

router = APIRouter(prefix="/api/hospital-ratings", tags=["hospital-ratings"])


class SubmitDetail(BaseModel):
    indicator_id: int
    actual_value: Optional[str] = None
    remark: Optional[str] = None


class SubmitBody(BaseModel):
    department_id: Optional[int] = None
    rating_cycle: str
    details: List[SubmitDetail]
    status: str = "submitted"  # "draft" or "submitted"


def _build_item_dict(item, ind, cat):
    return {
        "id": item.id,
        "indicator_id": ind.id,
        "name": ind.name,
        "category_name": cat.name if cat else None,
        "standard_value": ind.standard_value,
        "unit": ind.unit,
        "indicator_type": ind.indicator_type,
        "weight": float(ind.weight) if ind.weight else None,
        "actual_value": item.actual_value,
        "is_compliant": item.is_compliant,
        "score": item.score,
        "remark": item.gap_note,
    }


# ---------- Dashboard (delegates to existing /api/dashboard/director) ----------
# Frontend calls GET /api/hospital-ratings/dashboard which the API client maps.
# Actual logic lives in dashboard.py; this is kept as a thin passthrough.


# ---------- Submit ----------

@router.post("/submit")
def submit_rating(
    body: SubmitBody,
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
    ut=Depends(get_current_user_tenant),
):
    dept_id = body.department_id or ut.dept_id
    if not dept_id:
        raise HTTPException(400, "No department context")

    hg_set = db.query(StandardSet).filter(StandardSet.type == "hospital_grade").first()
    if not hg_set:
        raise HTTPException(400, "No hospital grade standard set found")

    dept = db.get(Department, dept_id)
    if not dept:
        raise HTTPException(404, "Department not found")

    a = Assessment(
        tenant_id=tenant_id,
        name=f"{dept.name} - {body.rating_cycle}三甲评级",
        target_level=1,
        department_id=dept_id,
        rating_cycle=body.rating_cycle,
        submitter_id=user.id,
        set_id=hg_set.id,
    )
    db.add(a)
    db.flush()

    items_out = []
    for d in body.details:
        ind = db.get(StdIndicator, d.indicator_id)
        if not ind:
            continue

        is_compliant = None
        score = None
        if d.actual_value is not None and ind.standard_value and ind.indicator_type:
            result = check_compliance(d.actual_value, ind.standard_value, ind.indicator_type)
            is_compliant = result["is_compliant"]
            score = result["score"]

        item = AssessmentItem(
            assessment_id=a.id,
            indicator_id=d.indicator_id,
            actual_value=d.actual_value,
            is_compliant=is_compliant,
            score=score,
            gap_note=d.remark,
            updated_at=datetime.now(timezone.utc) if d.actual_value else None,
        )
        db.add(item)
        db.flush()

        cat = db.get(StdCategory, ind.category_id)
        items_out.append(_build_item_dict(item, ind, cat))

    scored = [it for it in items_out if it["score"] is not None and it["weight"]]
    if scored:
        tw = sum(s["score"] * s["weight"] / 100 for s in scored)
        total_w = sum(s["weight"] for s in scored)
        a.total_score = round(tw / total_w * 100, 2) if total_w else 0

    a.status = body.status
    if body.status == "submitted":
        a.submitted_at = datetime.now(timezone.utc)
    _log(db, user.id, tenant_id,
         "submit" if body.status == "submitted" else "draft",
         "assessment", a.id, f"{dept.name} {body.rating_cycle} score:{a.total_score}")
    db.commit()

    return {
        "assessment_id": a.id,
        "total_score": float(a.total_score) if a.total_score else 0,
        "total_items": len(items_out),
        "compliant_count": sum(1 for it in items_out if it["is_compliant"]),
        "non_compliant_count": sum(1 for it in items_out if it["is_compliant"] is False),
        "items": items_out,
    }


# ---------- Edit & Resubmit ----------

@router.put("/submit/{aid}")
def update_and_resubmit(
    aid: int,
    body: SubmitBody,
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
    _=Depends(get_current_user_tenant),
):
    """科室负责人修改已退回的评级数据并重新提交"""
    a = db.query(Assessment).filter(
        Assessment.id == aid, Assessment.tenant_id == tenant_id
    ).first()
    if not a:
        raise HTTPException(404, "Assessment not found")
    if a.status not in ("draft", "rejected", "revising"):
        raise HTTPException(400, f"Cannot edit: current status is '{a.status}'")

    # Remove existing items
    db.query(AssessmentItem).filter(AssessmentItem.assessment_id == aid).delete()
    db.flush()

    items_out = []
    for d in body.details:
        ind = db.get(StdIndicator, d.indicator_id)
        if not ind:
            continue
        is_compliant = None
        score = None
        if d.actual_value is not None and ind.standard_value and ind.indicator_type:
            result = check_compliance(d.actual_value, ind.standard_value, ind.indicator_type)
            is_compliant = result["is_compliant"]
            score = result["score"]
        item = AssessmentItem(
            assessment_id=a.id, indicator_id=d.indicator_id,
            actual_value=d.actual_value, is_compliant=is_compliant,
            score=score, gap_note=d.remark,
            updated_at=datetime.now(timezone.utc) if d.actual_value else None,
        )
        db.add(item)
        db.flush()
        cat = db.get(StdCategory, ind.category_id)
        items_out.append(_build_item_dict(item, ind, cat))

    scored = [it for it in items_out if it["score"] is not None and it["weight"]]
    if scored:
        tw = sum(s["score"] * s["weight"] / 100 for s in scored)
        total_w = sum(s["weight"] for s in scored)
        a.total_score = round(tw / total_w * 100, 2) if total_w else 0

    a.status = body.status
    if body.status == "submitted":
        a.submitted_at = datetime.now(timezone.utc)
    _log(db, user.id, tenant_id,
         "edit" if body.status != "draft" else "draft",
         "assessment", a.id, f"resubmit score:{a.total_score}")
    db.commit()

    return {
        "assessment_id": a.id,
        "total_score": float(a.total_score) if a.total_score else 0,
        "total_items": len(items_out),
        "compliant_count": sum(1 for it in items_out if it["is_compliant"]),
        "non_compliant_count": sum(1 for it in items_out if it["is_compliant"] is False),
        "items": items_out,
    }


# ---------- My Department ----------

@router.get("/my-department")
def my_department_ratings(
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
    ut: UserTenant = Depends(get_current_user_tenant),
):
    hg_set = db.query(StandardSet).filter(StandardSet.type == "hospital_grade").first()
    set_id = hg_set.id if hg_set else None

    q = db.query(Assessment).filter(
        Assessment.tenant_id == tenant_id,
        Assessment.department_id == ut.dept_id,
    )
    if set_id:
        q = q.filter(Assessment.set_id == set_id)

    assessments = q.order_by(Assessment.created_at.desc()).all()
    return [
        {
            "id": a.id,
            "name": a.name,
            "rating_cycle": a.rating_cycle,
            "total_score": float(a.total_score) if a.total_score else None,
            "status": a.status,
            "submitted_at": a.submitted_at.isoformat() if a.submitted_at else None,
            "non_compliant_count": sum(1 for it in a.items if it.is_compliant is False),
            "total_items": len(a.items),
        }
        for a in assessments
    ]


# ---------- Report ----------

@router.get("/report/{aid}")
def get_report(
    aid: int,
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
):
    a = db.query(Assessment).filter(
        Assessment.id == aid, Assessment.tenant_id == tenant_id
    ).first()
    if not a:
        raise HTTPException(404, "Assessment not found")

    items = []
    for item in a.items:
        ind = db.get(StdIndicator, item.indicator_id)
        cat = db.get(StdCategory, ind.category_id) if ind else None
        items.append({
            "id": item.id,
            "indicator_id": item.indicator_id,
            "name": ind.name if ind else None,
            "category_name": cat.name if cat else None,
            "standard_value": ind.standard_value if ind else None,
            "unit": ind.unit if ind else None,
            "weight": float(ind.weight) if ind and ind.weight else None,
            "actual_value": item.actual_value,
            "is_compliant": item.is_compliant,
            "score": item.score,
            "remark": item.gap_note,
        })

    compliant = [i for i in items if i["is_compliant"]]
    # Per-category breakdown
    cat_stats = {}
    for i in items:
        cn = i["category_name"] or "未分类"
        if cn not in cat_stats:
            cat_stats[cn] = {"total": 0, "compliant": 0}
        cat_stats[cn]["total"] += 1
        if i["is_compliant"]:
            cat_stats[cn]["compliant"] += 1
    categories_breakdown = [
        {"name": cn, "total": s["total"], "compliant": s["compliant"],
         "rate": round(s["compliant"] / s["total"] * 100, 1) if s["total"] else 0}
        for cn, s in cat_stats.items()
    ]
    reviews = [
        {
            "id": r.id,
            "reviewer_id": r.reviewer_id,
            "action": r.action,
            "feedback": r.feedback,
            "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
        }
        for r in (a.reviews or [])
    ]
    return {
        "assessment_id": a.id,
        "name": a.name,
        "rating_cycle": a.rating_cycle,
        "total_score": float(a.total_score) if a.total_score else 0,
        "total_items": len(items),
        "compliant_count": len(compliant),
        "non_compliant_count": len(items) - len(compliant),
        "compliance_rate": f"{(len(compliant) / len(items) * 100):.1f}%" if items else "0%",
        "passed": float(a.total_score or 0) >= 60,
        "status": a.status,
        "items": items,
        "categories_breakdown": categories_breakdown,
        "reviews": reviews,
    }


# ---------- Standards tree ----------

@router.get("/standards")
def list_hospital_standards(db: Session = Depends(get_db)):
    categories = db.query(StdCategory).filter(
        StdCategory.parent_id.is_(None)
    ).order_by(StdCategory.sort_order).all()

    def _tree(cat):
        return {
            "id": cat.id,
            "name": cat.name,
            "code": cat.code,
            "weight": float(cat.weight) if cat.weight else None,
            "sort_order": cat.sort_order,
            "children": [_tree(c) for c in cat.children] if cat.children else None,
            "indicators": [
                {
                    "id": ind.id,
                    "code": ind.code,
                    "name": ind.name,
                    "standard_value": ind.standard_value,
                    "unit": ind.unit,
                    "indicator_type": ind.indicator_type,
                    "weight": float(ind.weight) if ind.weight else None,
                    "max_score": ind.max_score,
                }
                for ind in cat.indicators
            ],
        }

    return [_tree(c) for c in categories]

# ---------- PDF Report ----------

@router.get("/report/{aid}/pdf")
def download_pdf(
    aid: int,
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
):
    a = db.query(Assessment).filter(
        Assessment.id == aid, Assessment.tenant_id == tenant_id
    ).first()
    if not a:
        raise HTTPException(404, "Assessment not found")

    items = []
    for item in a.items:
        ind = db.get(StdIndicator, item.indicator_id)
        cat = db.get(StdCategory, ind.category_id) if ind else None
        items.append({
            "name": ind.name if ind else "?",
            "category": cat.name if cat else "?",
            "standard": (ind.standard_value or "") + (ind.unit or ""),
            "actual": item.actual_value or "-",
            "ok": item.is_compliant,
            "score": item.score or 0,
        })

    compliant = sum(1 for i in items if i["ok"])
    total = len(items)

    html = f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">
<style>
body{{font-family:"Microsoft YaHei",sans-serif;padding:40px;color:#1e293b}}
h1{{text-align:center;font-size:22px;margin-bottom:4px}}
.sub{{text-align:center;color:#64748b;font-size:13px;margin-bottom:20px}}
.stats{{display:flex;gap:12px;margin-bottom:20px}}
.stat{{flex:1;text-align:center;padding:12px;border-radius:6px;background:#f1f5f9}}
.stat .v{{font-size:24px;font-weight:700}}
.stat .l{{font-size:11px;color:#64748b}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#f1f5f9;padding:8px;text-align:left;border-bottom:2px solid #e2e8f0}}
td{{padding:8px;border-bottom:1px solid #e2e8f0}}
.pass{{color:#16a34a;font-weight:600}}
.fail{{color:#dc2626;font-weight:600}}
</style></head><body>
<h1>{a.name}</h1>
<p class="sub">评级周期: {a.rating_cycle or '-'} | 状态: {a.status}</p>
<div class="stats">
<div class="stat"><div class="v" style="color:{'#16a34a' if float(a.total_score or 0)>=60 else '#dc2626'}">{a.total_score or 0}</div><div class="l">总分</div></div>
<div class="stat"><div class="v" style="color:#3b82f6">{len(compliant) if isinstance(compliant, list) else compliant}/{total}</div><div class="l">达标率</div></div>
<div class="stat"><div class="v">{total}</div><div class="l">总指标</div></div>
</div>
<table><tr><th>分类</th><th>指标</th><th>标准值</th><th>实际值</th><th>结果</th></tr>
"""
    for i in items:
        cls = "pass" if i["ok"] else "fail"
        tag = "达标" if i["ok"] else "未达标"
        html += f"<tr><td>{i['category']}</td><td>{i['name']}</td><td>{i['standard']}</td><td style='{cls}'>{i['actual']}</td><td class='{cls}'>{tag}</td></tr>"
    html += "</table></body></html>"

    try:
        from weasyprint import HTML
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        buf = BytesIO()
        HTML(string=html).write_pdf(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename=report_{aid}.pdf"})
    except Exception:
        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html)


# ---------- Excel Import Assessment Data ----------

@router.post("/import-data")
def import_assessment_data(
    file: UploadFile = File(...),
    rating_cycle: str = "2025年度",
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
    ut=Depends(get_current_user_tenant),
):
    """Excel 批量导入科室评估数据。列: indicator_code | actual_value | remark"""
    dept_id = ut.dept_id
    if not dept_id:
        raise HTTPException(400, "No department context")

    hg_set = db.query(StandardSet).filter(StandardSet.type == "hospital_grade").first()
    if not hg_set:
        raise HTTPException(400, "No hospital grade standard set found")

    dept = db.get(Department, dept_id)
    if not dept:
        raise HTTPException(404, "Department not found")

    # Parse Excel
    import tempfile, os
    suffix = os.path.splitext(file.filename or "data.xlsx")[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    finally:
        os.unlink(tmp_path)

    # Create assessment
    a = Assessment(
        tenant_id=tenant_id,
        name=f"{dept.name} - {rating_cycle}三甲评级",
        target_level=1, department_id=dept_id,
        rating_cycle=rating_cycle, submitter_id=user.id,
        set_id=hg_set.id, status="draft",
    )
    db.add(a)
    db.flush()

    count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        actual = str(row[1]).strip() if len(row) > 1 and row[1] else None
        remark = str(row[2]).strip() if len(row) > 2 and row[2] else None

        ind = db.query(StdIndicator).filter(StdIndicator.code == code).first()
        if not ind:
            continue

        is_compliant = None
        score = None
        if actual and ind.standard_value and ind.indicator_type:
            result = check_compliance(actual, ind.standard_value, ind.indicator_type)
            is_compliant = result["is_compliant"]
            score = result["score"]

        db.add(AssessmentItem(
            assessment_id=a.id, indicator_id=ind.id,
            actual_value=actual, is_compliant=is_compliant,
            score=score, gap_note=remark,
            updated_at=datetime.now(timezone.utc) if actual else None,
        ))
        count += 1

    # Calculate total score
    items = a.items
    tw, total_w = 0, 0
    for item in items:
        ind = db.get(StdIndicator, item.indicator_id)
        if ind and item.score is not None and ind.weight:
            tw += item.score * float(ind.weight) / 100
            total_w += float(ind.weight)
    a.total_score = round(tw / total_w * 100, 2) if total_w else 0

    _log(db, user.id, tenant_id, "import", "assessment", a.id, f"Excel import {count} items score:{a.total_score}")
    db.commit()

    return {"ok": True, "assessment_id": a.id, "count": count, "total_score": float(a.total_score) if a.total_score else 0}


# ---------- Copy from previous cycle ----------

@router.post("/copy-previous")
def copy_previous_cycle(
    rating_cycle: str = "2025年度",
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
    ut=Depends(get_current_user_tenant),
):
    """复制上一周期的评估数据作为新评估的起点"""
    dept_id = ut.dept_id
    if not dept_id:
        raise HTTPException(400, "No department context")

    prev = db.query(Assessment).filter(
        Assessment.tenant_id == tenant_id,
        Assessment.department_id == dept_id,
    ).order_by(Assessment.created_at.desc()).first()

    if not prev:
        raise HTTPException(404, "No previous assessment to copy")

    hg_set = db.query(StandardSet).filter(StandardSet.type == "hospital_grade").first()
    dept = db.get(Department, dept_id)

    a = Assessment(
        tenant_id=tenant_id,
        name=f"{dept.name} - {rating_cycle}三甲评级",
        target_level=1, department_id=dept_id,
        rating_cycle=rating_cycle, submitter_id=user.id,
        set_id=hg_set.id, status="draft",
    )
    db.add(a)
    db.flush()

    copied = 0
    for prev_item in prev.items:
        ind = db.get(StdIndicator, prev_item.indicator_id)
        if not ind:
            continue
        actual = prev_item.actual_value
        is_compliant = None
        score = None
        if actual and ind.standard_value and ind.indicator_type:
            result = check_compliance(actual, ind.standard_value, ind.indicator_type)
            is_compliant = result["is_compliant"]
            score = result["score"]
        db.add(AssessmentItem(
            assessment_id=a.id, indicator_id=ind.id,
            actual_value=actual, is_compliant=is_compliant,
            score=score, gap_note=prev_item.gap_note,
        ))
        copied += 1

    db.commit()
    return {"ok": True, "assessment_id": a.id, "copied": copied}


# ---------- History Comparison ----------

@router.get("/compare/{dept_id}")
def compare_history(
    dept_id: int,
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
):
    """对比同一科室不同周期的评级数据"""
    assessments = db.query(Assessment).filter(
        Assessment.tenant_id == tenant_id,
        Assessment.department_id == dept_id,
    ).order_by(Assessment.created_at.desc()).limit(5).all()

    result = []
    for a in assessments:
        compliant = sum(1 for it in a.items if it.is_compliant)
        result.append({
            "id": a.id, "name": a.name, "rating_cycle": a.rating_cycle,
            "total_score": float(a.total_score) if a.total_score else 0,
            "compliant": compliant, "total": len(a.items),
            "status": a.status,
            "submitted_at": a.submitted_at.isoformat() if a.submitted_at else None,
        })
    return result
